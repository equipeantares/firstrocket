#Equipe Antares - UNICAMP

#This code read data from BMP180 sensor and store it in an excel file
#Must install openpyxl library
#Use pip install openpyxl on terminal

import time
import Adafruit_BMP.BMP085 as BMP085
from openpyxl import Workbook

wb = Workbook()

#define excel file name
dest_filename = 'Rocket Data.xlsx'

#select the active sheet and change its name
ws1 = wb.active
ws1.title = "Rocket Data"

sensor = BMP085.BMP085()

#save the initial time
time_base = time.time()

#save the initial altitude
alt_base = sensor.read_altitude()

#define the column names
ws1.cell(column = 1, row = 1, value = "Time")
ws1.cell(column = 2, row = 1, value = "Altitude")
ws1.cell(column = 3, row = 1, value = "Temperature")
ws1.cell(column = 4, row = 1, value = "Pressure")

#initial row for data
row = 2

#colect data from bmp180 and save in the excel file
while row < 10:
	ws1.cell(column = 1, row = row, value = time.time() - time_base)
	ws1.cell(column = 2, row = row, value = sensor.read_altitude() - alt_base)
	ws1.cell(column = 3, row = row, value = sensor.read_temperature())
	ws1.cell(column = 4, row = row, value = sensor.read_pressure())
	row = row + 1
	time.sleep(1)

#save excel file
wb.save(filename = dest_filename)
